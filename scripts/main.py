import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import datetime

# Load Excel file
def load_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    return workbook

# Export Excel After Writing
def save_excel(workbook, file_path):
    workbook.save(file_path)

# Get Today if not present in sheet then User will select
def get_sheet(workbook):
    days = workbook.sheetnames
    today = datetime.datetime.now().strftime("%A")
    if today not in days:
        print(f"Today is {today}. No sheet found for today.")
        print("Select a day :")
        for i, day in enumerate(days):
            print(f"{i + 1}. {day}")
        while True:
            day_index = int(input("Enter the number of the day: ")) - 1
            if 0 <= day_index < len(days):
                return workbook[days[day_index]]
            else:
                print("Invalid day index, please try again.")
    else:
        print(f"Today is {today}. Found sheet for today.")
        return workbook[today]

# WebDriver Initialization
def setup_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")  # headless mode
    driver = webdriver.Chrome(options=options)
    driver.get("https://www.google.com/search?q=")
    time.sleep(1)
    return driver

# Search keywords and find suggestions
def process_keywords(sheet, driver):
    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=2, max_col=5):

        # Keywords start from 3rd row to max row count
        # Column 2nd for keywords, Column 3rd for the values
        # Column 3rd for longest option, Column 4th for shortest option to write

        keyword = row[1].value  # 3rd Column (Storing Keyword values to the keyword variable)
        if not keyword: # if keyword value is empty or None then skip that row
            continue

        #print(f"Processing '{keyword}'...")
        #print([cell.value for cell in row])

        try:
            search_box = driver.find_element(by=By.ID, value="APjFqb")
            search_box.clear()
            search_box.send_keys(keyword)
            time.sleep(0.7)

            # Extract suggestions
            #suggestions = driver.find_elements("xpath", "//ul[@role='listbox']//li//span") # with ancestor div
            suggestions = driver.find_elements("xpath", "//ul[@role='listbox']//li//span[not(ancestor::div[@class='ClJ9Yb'])]") # without ancestor div

            suggestion_texts = [suggestion.text for suggestion in suggestions if suggestion.text]
            #print([suggestion.text for suggestion in suggestions if suggestion.text])

            # Find longest and shortest options
            if suggestion_texts:
                longest_option = max(suggestion_texts, key=len)
                shortest_option = min(suggestion_texts, key=len)

                row[2].value = longest_option  # Column 4 for longest option
                row[3].value = shortest_option  # Column 5 for shortest option
                print(f"Processed '{keyword}': Longest='{longest_option}', Shortest='{shortest_option}'")
            else:
                print(f"No suggestions found for '{keyword}'")

            time.sleep(0.5)

        except Exception as e:
            print(f"Error processing '{keyword}': {e}")

def main():
    excel_path = "../input/eee.xlsx"
    output_path = "../output/search_done.xlsx"
    workbook = load_excel(excel_path)
    sheet = get_sheet(workbook)

    driver = setup_driver()

    try:
        process_keywords(sheet, driver)
    finally:
        driver.quit()

    save_excel(workbook, output_path)
    print(f"New excel file saved to {output_path}. Done!")

if __name__ == "__main__":
    main()